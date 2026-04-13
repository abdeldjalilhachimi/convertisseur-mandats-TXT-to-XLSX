#!/usr/bin/env python3
"""
Convertisseur TXT -> XLSX pour fichiers Mandats (format Algérie Poste / CCP).

Format attendu (62 caractères par ligne) :
  En-tête : * + 11 zéros + 10 ch. ref + 13 ch. montant total (×10) + 6 ch. nb lignes
            + 6 ch. MMAAAA + 8 ch. code organisme + 6 ch. n° lot + 1 ch.
  Détail  : * + 8 ch. préfixe + 12 ch. RIB/CCP + 13 ch. montant (centimes)
            + 27 ch. nom bénéficiaire + 1 ch. type

Utilisation :
  - GUI  : python3 txt_to_xlsx.py
  - CLI  : python3 txt_to_xlsx.py fichier.TXT [fichier2.TXT ...]
           python3 txt_to_xlsx.py --dir /chemin/dossier
"""
from __future__ import annotations
import argparse
import sys
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_MARKER = "*"
LINE_LEN = 62
AMOUNT_ALIGN = Alignment(horizontal="right")


def format_rib(rib: str) -> str:
    """CCP accounts (12 chars, starts with '000') → 'NNNNNNNNNN/KK'."""
    if len(rib) == 12 and rib[:3] == "000" and "/" not in rib:
        return f"{rib[:10]}/{rib[10:]}"
    return rib


def fmt_amount(value: float) -> str:
    """Format a DZD amount as '101 370.75' (space thousands, period decimal)."""
    return f"{value:,.2f}".replace(",", " ")


def parse_header(line: str) -> dict:
    """Parse la ligne d'en-tête — deux formats supportés :
    • Fixe (62 car.) : format original
    • Espaces        : *RIB F1 TOTAL NB_LIGNES MMYYYY ORGANISME F6 LOT FIN
    """
    if not line.startswith(HEADER_MARKER):
        raise ValueError(f"En-tête invalide : {line!r}")

    # ── Format espace-séparé (ligne > 62 car.) ───────────────────────────────
    # Layout fixe (9 tokens séparés par espaces) :
    #  [0] RIB compte   [1] ???    [2] réf/montant  [3] nb_lignes
    #  [4] MMYYYY       [5] code organisme  [6] ???  [7] lot  [8] fin
    rest = line[1:].rstrip()
    tokens = [t for t in rest.split(" ") if t]
    if len(tokens) >= 8 and len(rest) > LINE_LEN - 1:
        raw_periode = tokens[4] if len(tokens) > 4 else ""
        periode = (f"{raw_periode[:2]}/{raw_periode[2:]}"
                   if len(raw_periode) == 6 and raw_periode.isdigit() else raw_periode)
        return {
            "rib_compte":    tokens[0],
            "reference":     tokens[0],
            "montant_total": 0.0,          # calculé depuis les détails
            "nb_lignes":     int(tokens[3]) if tokens[3].isdigit() else 0,
            "periode":       periode,
            "organisme":     tokens[5] if len(tokens) > 5 else "",
            "lot":           tokens[7].strip() if len(tokens) > 7 else "",
            "fin":           tokens[8] if len(tokens) > 8 else "",
            "format":        "space",
        }

    # ── Format fixe original (62 car.) ───────────────────────────────────────
    if len(line) < LINE_LEN:
        raise ValueError(f"En-tête invalide (trop courte) : {line!r}")
    return {
        "rib_compte":   "",
        "reference":    line[12:22],
        "montant_total": int(line[22:35]) / 1000,
        "nb_lignes":    int(line[35:41]),
        "periode":      f"{line[41:43]}/{line[43:47]}",
        "organisme":    line[47:55].strip(),
        "lot":          line[55:61].strip(),
        "fin":          line[61:62],
        "format":       "fixed",
    }


def _parse_detail_space(tokens: list[str], num: int,
                        prefixe: str, rib: str) -> dict:
    """Parse le reste d'une ligne détail en format espace-séparé."""
    idx = 0
    # Saute le champ '000000' (remplissage)
    if idx < len(tokens) and tokens[idx] == "000000":
        idx += 1
    # Montant (peut finir par '.')
    amount_str = tokens[idx].rstrip(".") if idx < len(tokens) else "0"
    montant = int(amount_str) / 100
    idx += 1
    # Type = dernier token à 1 chiffre
    type_ = ""
    name_end = len(tokens)
    if tokens and tokens[-1].isdigit() and len(tokens[-1]) == 1:
        type_ = tokens[-1]
        name_end -= 1
    beneficiaire = " ".join(tokens[idx:name_end]).strip()
    return {
        "n": num, "prefixe": prefixe, "rib": rib,
        "montant": montant, "beneficiaire": beneficiaire, "type": type_,
    }


def parse_detail(line: str, num: int) -> dict:
    """Parse une ligne détail — trois formats supportés :
    • Fixe (62 car.)   : format original
    • Espace-montant   : *RIB12 · 000000 · AMOUNT · NAME · TYPE
    • Espace-CCP       : * RIB10 · KEY · 000000 · AMOUNT · NAME · TYPE
    """
    if not line.startswith(HEADER_MARKER):
        raise ValueError(f"Ligne {num} invalide : {line!r}")
    s = line.rstrip()

    # ── Format Espace-CCP : '* ' en début (espace après le marqueur) ─────────
    if len(s) > 1 and s[1] == " ":
        content = s[2:].lstrip()                # après '* '
        prefixe  = content[0:8]
        rib_base = content[8:18]                # RIB 10 chiffres
        tail     = content[18:].strip()
        tokens   = [t for t in tail.split(" ") if t]
        idx = 0
        # Clé de contrôle (2 chiffres) → concaténée au RIB : '0004807758/56'
        if tokens and tokens[idx].isdigit() and len(tokens[idx]) == 2:
            rib = f"{rib_base}/{tokens[idx]}"
            idx += 1
        else:
            rib = rib_base
        return _parse_detail_space(tokens[idx:], num, prefixe, rib)

    # ── Format Espace-montant : espace en position 21 ────────────────────────
    if len(s) > 21 and s[21] == " ":
        prefixe = s[1:9]
        rib     = s[9:21]
        tokens  = [t for t in s[21:].split(" ") if t]
        return _parse_detail_space(tokens, num, prefixe, rib)

    # ── Format fixe original (62 car.) ───────────────────────────────────────
    s = s.ljust(LINE_LEN)
    return {
        "n":            num,
        "prefixe":      s[1:9],
        "rib":          format_rib(s[9:21]),
        "montant":      int(s[21:34]) / 100,
        "beneficiaire": s[34:61].rstrip(),
        "type":         s[61:62],
    }


def parse_file(path: Path) -> tuple[dict, list[dict]]:
    raw = path.read_text(encoding="utf-8", errors="replace").splitlines()
    lines = [l.rstrip("\r\n ") for l in raw if l.strip()]
    if not lines:
        raise ValueError(f"Fichier vide : {path}")
    header = parse_header(lines[0])
    details = [parse_detail(l, i) for i, l in enumerate(lines[1:], start=1)]
    return header, details


def add_details_sheet(wb, files_data: list) -> None:
    """Add a 'Détails' sheet: per-file header block followed by that file's mandate rows."""
    ws = wb.create_sheet("Détails")

    FILE_FILL  = PatternFill("solid", fgColor="305496")
    FILE_FONT  = Font(bold=True, color="FFFFFF", size=12)
    COL_FILL   = PatternFill("solid", fgColor="BDD7EE")
    COL_FONT   = Font(bold=True)
    META_LABEL = Font(bold=True, color="305496")
    COL_HEADERS = ["N°", "Préfixe", "RIB / CCP", "Montant (DZD)", "Bénéficiaire", "Type"]
    N_COLS = len(COL_HEADERS)

    current_row = 1

    for header, details, src in files_data:
        # ── File title bar ──────────────────────────────────────────────────
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=N_COLS)
        title_cell = ws.cell(row=current_row, column=1,
                             value=f"Fichier : {src.name}")
        title_cell.font = FILE_FONT
        title_cell.fill = FILE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # ── Metadata key-value pairs ─────────────────────────────────────────
        meta = []
        if header.get("rib_compte"):
            meta.append(("RIB / CCP organisme", header["rib_compte"]))
        else:
            meta.append(("Référence", header["reference"]))
        meta += [
            ("Période (MM/AAAA)",      header["periode"]),
            ("Code organisme",         header["organisme"]),
            ("N° lot",                 header["lot"]),
            ("Lignes déclarées",       header["nb_lignes"]),
            ("Lignes lues",            len(details)),
            ("Montant déclaré (DZD)",  fmt_amount(header["montant_total"])),
            ("Montant calculé (DZD)",  fmt_amount(sum(d["montant"] for d in details))),
        ]
        for label, value in meta:
            lbl = ws.cell(row=current_row, column=1, value=label)
            lbl.font = META_LABEL
            val = ws.cell(row=current_row, column=2, value=value)
            if "Montant" in label:
                val.alignment = AMOUNT_ALIGN
            current_row += 1

        current_row += 1  # blank separator before table

        # ── Table header ─────────────────────────────────────────────────────
        for col, h in enumerate(COL_HEADERS, start=1):
            cell = ws.cell(row=current_row, column=col, value=h)
            cell.font = COL_FONT
            cell.fill = COL_FILL
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # ── Mandate rows ──────────────────────────────────────────────────────
        for d in details:
            ws.cell(row=current_row, column=1, value=d["n"])
            ws.cell(row=current_row, column=2, value=d["prefixe"])
            rib = ws.cell(row=current_row, column=3, value=d["rib"])
            rib.number_format = "@"
            amt = ws.cell(row=current_row, column=4, value=fmt_amount(d["montant"]))
            amt.alignment = AMOUNT_ALIGN
            ws.cell(row=current_row, column=5, value=d["beneficiaire"])
            ws.cell(row=current_row, column=6, value=d["type"])
            current_row += 1

        # ── Subtotal ──────────────────────────────────────────────────────────
        ws.cell(row=current_row, column=1, value="TOTAL").font = Font(bold=True)
        sub = ws.cell(row=current_row, column=4,
                      value=fmt_amount(sum(d["montant"] for d in details)))
        sub.font = Font(bold=True)
        sub.alignment = AMOUNT_ALIGN
        current_row += 2  # blank row between file sections

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [28, 24, 16, 15, 32, 6]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_xlsx(header: dict, details: list[dict], src: Path, out: Path) -> None:
    wb = Workbook()

    # --- Feuille En-tête ---
    ws_h = wb.active
    ws_h.title = "En-tête"
    ws_h["A1"] = "Champ"
    ws_h["B1"] = "Valeur"
    for c in ("A1", "B1"):
        ws_h[c].font = Font(bold=True, color="FFFFFF")
        ws_h[c].fill = PatternFill("solid", fgColor="305496")
    rows_h = [("Fichier source", src.name)]
    if header.get("rib_compte"):
        rows_h.append(("RIB / CCP organisme", header["rib_compte"]))
    else:
        rows_h.append(("Référence", header["reference"]))
    rows_h += [
        ("Période (MM/AAAA)", header["periode"]),
        ("Code organisme",    header["organisme"]),
        ("N° lot",            header["lot"]),
        ("Lignes déclarées",  header["nb_lignes"]),
        ("Lignes lues",       len(details)),
        ("Montant total (DZD) déclaré", fmt_amount(header["montant_total"])),
        ("Montant total (DZD) calculé", fmt_amount(sum(d["montant"] for d in details))),
    ]
    for i, (k, v) in enumerate(rows_h, start=2):
        ws_h.cell(row=i, column=1, value=k)
        ws_h.cell(row=i, column=2, value=v)
    # right-align the two amount rows (last two rows)
    last = len(rows_h) + 1
    for row in (last - 1, last):
        ws_h.cell(row=row, column=2).alignment = AMOUNT_ALIGN
    ws_h.column_dimensions["A"].width = 32
    ws_h.column_dimensions["B"].width = 28

    # --- Feuille Mandats ---
    ws = wb.create_sheet("Mandats")
    headers = ["N°", "Préfixe", "RIB / CCP", "Montant (DZD)", "Bénéficiaire", "Type"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center")

    for r, d in enumerate(details, start=2):
        ws.cell(row=r, column=1, value=d["n"])
        ws.cell(row=r, column=2, value=d["prefixe"])
        rib_cell = ws.cell(row=r, column=3, value=d["rib"])
        rib_cell.number_format = "@"
        amt_cell = ws.cell(row=r, column=4, value=fmt_amount(d["montant"]))
        amt_cell.alignment = AMOUNT_ALIGN
        ws.cell(row=r, column=5, value=d["beneficiaire"])
        ws.cell(row=r, column=6, value=d["type"])

    # total
    total_row = len(details) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=4,
                         value=fmt_amount(sum(d["montant"] for d in details)))
    total_cell.font = Font(bold=True)
    total_cell.alignment = AMOUNT_ALIGN

    # largeurs
    widths = [6, 10, 16, 15, 32, 6]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{total_row - 1}"

    # --- Feuille Détails ---
    add_details_sheet(wb, [(header, details, src)])

    wb.save(out)


def convert(path: Path) -> Path:
    header, details = parse_file(path)
    out = path.with_suffix(".xlsx")
    write_xlsx(header, details, path, out)
    return out


def convert_many(paths: Iterable[Path]) -> list[tuple[Path, Path | Exception]]:
    results = []
    for p in paths:
        try:
            results.append((p, convert(p)))
        except Exception as e:
            results.append((p, e))
    return results


# --------------------- GUI ---------------------

def run_gui(preloaded: list[Path] | None = None):
    import tkinter as tk
    from tkinter import filedialog, messagebox, scrolledtext, ttk
    import subprocess

    # Try drag-drop support
    try:
        from tkinterdnd2 import TkinterDnD, DND_FILES
        root = TkinterDnD.Tk()
        dnd_ok = True
    except Exception:
        root = tk.Tk()
        dnd_ok = False

    root.title("TXT → XLSX (Mandats)")
    root.geometry("720x560")
    root.configure(bg="#f5f6fa")

    # Header
    tk.Label(root, text="Convertisseur Mandats TXT → XLSX",
             font=("Helvetica", 18, "bold"), bg="#f5f6fa",
             fg="#1f3a5f").pack(pady=(16, 4))
    tk.Label(root,
             text="Glissez vos fichiers .TXT ici, ou cliquez pour parcourir",
             font=("Helvetica", 12), bg="#f5f6fa", fg="#555").pack(pady=(0, 10))

    # Drop zone
    drop = tk.Frame(root, bg="#ffffff",
                    highlightbackground="#4a90e2",
                    highlightthickness=2, bd=0)
    drop.pack(fill="x", padx=20, pady=6, ipady=30)

    drop_label = tk.Label(
        drop,
        text=("📄  Déposez le(s) fichier(s) ici\n\nou cliquez pour choisir"
              if dnd_ok else "📄  Cliquez pour choisir le(s) fichier(s)"),
        font=("Helvetica", 14), bg="#ffffff", fg="#4a90e2",
        cursor="hand2", justify="center",
    )
    drop_label.pack(expand=True, fill="both")

    # Log
    log_frame = tk.Frame(root, bg="#f5f6fa")
    log_frame.pack(fill="both", expand=True, padx=20, pady=(10, 4))
    tk.Label(log_frame, text="Résultats :", font=("Helvetica", 11, "bold"),
             bg="#f5f6fa", anchor="w").pack(fill="x")
    log = scrolledtext.ScrolledText(log_frame, height=10,
                                    font=("Menlo", 11), bg="#ffffff")
    log.pack(fill="both", expand=True)

    last_outputs: list[Path] = []

    def write(msg: str, color: str | None = None):
        log.insert("end", msg + "\n")
        log.see("end")
        root.update()

    def process(paths: list[Path]):
        if not paths:
            return
        # expand folders
        expanded: list[Path] = []
        for p in paths:
            if p.is_dir():
                expanded.extend(sorted(p.glob("*.TXT")))
                expanded.extend(sorted(p.glob("*.txt")))
            elif p.is_file():
                expanded.append(p)
        expanded = [p for p in expanded if p.exists()]
        if not expanded:
            messagebox.showwarning("Rien à convertir",
                                   "Aucun fichier .TXT valide.")
            return
        write(f"\n— Conversion de {len(expanded)} fichier(s) —")
        results = convert_many(expanded)
        ok_paths: list[Path] = []
        for src, res in results:
            if isinstance(res, Path):
                write(f"✓ {src.name} → {res.name}")
                ok_paths.append(res)
            else:
                write(f"✗ {src.name} : {res}")
        last_outputs.clear()
        last_outputs.extend(ok_paths)
        btn_open.config(state=("normal" if ok_paths else "disabled"))
        btn_reveal.config(state=("normal" if ok_paths else "disabled"))
        messagebox.showinfo(
            "Terminé",
            f"{len(ok_paths)}/{len(expanded)} fichier(s) convertis.",
        )

    def browse(_=None):
        paths = filedialog.askopenfilenames(
            title="Choisir fichier(s) TXT",
            filetypes=[("Fichiers TXT", "*.TXT *.txt"), ("Tous", "*.*")],
        )
        if paths:
            process([Path(p) for p in paths])

    def browse_folder():
        d = filedialog.askdirectory(title="Choisir un dossier")
        if d:
            process([Path(d)])

    def on_drop(event):
        # event.data = liste de chemins séparés par espaces, entourés d'accolades si espaces
        raw = event.data
        paths = []
        buf, in_brace = "", False
        for ch in raw:
            if ch == "{":
                in_brace = True
            elif ch == "}":
                in_brace = False
                if buf:
                    paths.append(buf)
                    buf = ""
            elif ch == " " and not in_brace:
                if buf:
                    paths.append(buf)
                    buf = ""
            else:
                buf += ch
        if buf:
            paths.append(buf)
        process([Path(p) for p in paths])

    drop_label.bind("<Button-1>", browse)
    drop.bind("<Button-1>", browse)

    if dnd_ok:
        drop.drop_target_register(DND_FILES)
        drop.dnd_bind("<<Drop>>", on_drop)
        drop_label.drop_target_register(DND_FILES)
        drop_label.dnd_bind("<<Drop>>", on_drop)

    # Boutons d'actions
    btns = tk.Frame(root, bg="#f5f6fa")
    btns.pack(pady=10)

    def open_outputs():
        for p in last_outputs:
            subprocess.run(["open", str(p)])

    def reveal_outputs():
        for p in last_outputs:
            subprocess.run(["open", "-R", str(p)])

    tk.Button(btns, text="Parcourir fichiers…", command=browse,
              width=18).pack(side="left", padx=4)
    tk.Button(btns, text="Parcourir dossier…", command=browse_folder,
              width=18).pack(side="left", padx=4)
    btn_open = tk.Button(btns, text="Ouvrir XLSX", command=open_outputs,
                         width=14, state="disabled")
    btn_open.pack(side="left", padx=4)
    btn_reveal = tk.Button(btns, text="Révéler dans Finder",
                           command=reveal_outputs, width=18, state="disabled")
    btn_reveal.pack(side="left", padx=4)

    # Fichiers pré-chargés (drop sur l'icône du .app)
    if preloaded:
        root.after(200, lambda: process(preloaded))

    root.mainloop()


# --------------------- CLI ---------------------

def main():
    ap = argparse.ArgumentParser(description="Convertit fichiers Mandats TXT en XLSX.")
    ap.add_argument("files", nargs="*", help="Fichier(s) TXT")
    ap.add_argument("--dir", help="Dossier : convertit tous les .TXT")
    ap.add_argument("--gui", action="store_true", help="Lance l'interface graphique")
    args = ap.parse_args()

    paths: list[Path] = [Path(f) for f in args.files]
    if args.dir:
        d = Path(args.dir)
        paths.extend(sorted(d.glob("*.TXT")))
        paths.extend(sorted(d.glob("*.txt")))

    if args.gui:
        run_gui(preloaded=paths or None)
        return

    if not paths:
        run_gui()
        return

    fail = 0
    for src, res in convert_many(paths):
        if isinstance(res, Path):
            print(f"OK  {src} -> {res}")
        else:
            print(f"ERR {src} : {res}", file=sys.stderr)
            fail += 1
    sys.exit(1 if fail else 0)


if __name__ == "__main__":
    main()
