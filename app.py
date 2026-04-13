import io
import tempfile
import zipfile
from pathlib import Path

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from txt_to_xlsx import convert, parse_file, fmt_amount, add_details_sheet

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
BLUE_FILL = PatternFill("solid", fgColor="305496")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
AMOUNT_ALIGN = Alignment(horizontal="right")


def write_xlsx_merged(files_data: list, buf: io.BytesIO) -> None:
    """Write one XLSX combining all files. files_data = [(header, details, src_path), ...]"""
    wb = Workbook()

    # --- Summary sheet ---
    ws_h = wb.active
    ws_h.title = "En-tête"
    summary_cols = [
        "Fichier", "Référence", "Période", "Organisme", "N° lot",
        "Lignes déclarées", "Détails lus", "Montant déclaré (DZD)", "Montant calculé (DZD)",
    ]
    ws_h.append(summary_cols)
    for col in range(1, len(summary_cols) + 1):
        cell = ws_h.cell(row=1, column=col)
        cell.font = WHITE_BOLD
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center")

    for h, details, src in files_data:
        row_data = [
            src.name,
            h["reference"],
            h["periode"],
            h["organisme"],
            h["lot"],
            h["nb_lignes"],
            len(details),
            fmt_amount(h["montant_total"]),
            fmt_amount(sum(d["montant"] for d in details)),
        ]
        ws_h.append(row_data)

    for row in ws_h.iter_rows(min_row=2, max_row=ws_h.max_row, min_col=8, max_col=9):
        for cell in row:
            cell.alignment = AMOUNT_ALIGN

    col_widths = [28, 14, 12, 12, 8, 18, 12, 24, 24]
    for i, w in enumerate(col_widths, start=1):
        ws_h.column_dimensions[get_column_letter(i)].width = w

    # --- Merged mandats sheet ---
    ws = wb.create_sheet("Mandats")
    headers = ["Fichier", "N°", "Préfixe", "RIB / CCP", "Montant (DZD)", "Bénéficiaire", "Type"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = WHITE_BOLD
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    grand_total = 0.0
    for h, details, src in files_data:
        for d in details:
            ws.cell(row=row_num, column=1, value=src.name)
            ws.cell(row=row_num, column=2, value=d["n"])
            ws.cell(row=row_num, column=3, value=d["prefixe"])
            rib_cell = ws.cell(row=row_num, column=4, value=d["rib"])
            rib_cell.number_format = "@"
            amt_cell = ws.cell(row=row_num, column=5, value=fmt_amount(d["montant"]))
            amt_cell.alignment = AMOUNT_ALIGN
            ws.cell(row=row_num, column=6, value=d["beneficiaire"])
            ws.cell(row=row_num, column=7, value=d["type"])
            grand_total += d["montant"]
            row_num += 1

    total_row = row_num
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=5, value=fmt_amount(grand_total))
    total_cell.font = Font(bold=True)
    total_cell.alignment = AMOUNT_ALIGN

    col_widths_m = [28, 6, 10, 16, 15, 32, 6]
    for i, w in enumerate(col_widths_m, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{total_row - 1}"

    add_details_sheet(wb, files_data)

    wb.save(buf)


# ─── UI ───────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Mandats TXT → XLSX", page_icon="📄", layout="centered")

st.title("Convertisseur Mandats TXT → XLSX")

uploaded_files = st.file_uploader(
    label="Fichiers .TXT",
    type=["txt", "TXT"],
    accept_multiple_files=True,
    label_visibility="collapsed",
)

if not uploaded_files:
    st.stop()

mode = st.radio(
    "Mode de conversion",
    options=["Fichiers séparés", "Fusionner en un seul fichier"],
    horizontal=True,
)

if not st.button("Convertir", type="primary", use_container_width=True):
    st.stop()

with tempfile.TemporaryDirectory() as tmp:
    tmp_path = Path(tmp)

    # Write all uploaded files to temp dir
    src_paths = []
    for f in uploaded_files:
        p = tmp_path / f.name
        p.write_bytes(f.read())
        src_paths.append(p)

    if mode == "Fichiers séparés":
        results = []
        errors = []

        with st.spinner("Conversion en cours…"):
            for src in src_paths:
                try:
                    out = convert(src)
                    results.append((src.name, out.read_bytes()))
                except Exception as e:
                    errors.append((src.name, str(e)))

        for name, err in errors:
            st.error(f"**{name}** — {err}")

        if results:
            if len(results) == 1:
                name, data = results[0]
                st.success(f"Converti : **{Path(name).stem}.xlsx**")
                st.download_button(
                    label="⬇️  Télécharger XLSX",
                    data=data,
                    file_name=Path(name).stem + ".xlsx",
                    mime=XLSX_MIME,
                    use_container_width=True,
                )
            else:
                # Bundle in a ZIP
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for name, data in results:
                        zf.writestr(Path(name).stem + ".xlsx", data)
                zip_buf.seek(0)
                st.success(f"{len(results)}/{len(src_paths)} fichier(s) converti(s).")
                st.download_button(
                    label="⬇️  Télécharger tous les XLSX (.zip)",
                    data=zip_buf,
                    file_name="mandats_xlsx.zip",
                    mime="application/zip",
                    use_container_width=True,
                )

    else:  # Merge
        parsed = []
        errors = []

        with st.spinner("Lecture et fusion en cours…"):
            for src in src_paths:
                try:
                    header, details = parse_file(src)
                    parsed.append((header, details, src))
                except Exception as e:
                    errors.append((src.name, str(e)))

        for name, err in errors:
            st.error(f"**{name}** — {err}")

        if parsed:
            buf = io.BytesIO()
            write_xlsx_merged(parsed, buf)
            buf.seek(0)
            st.success(f"{len(parsed)}/{len(src_paths)} fichier(s) fusionné(s).")
            st.download_button(
                label="⬇️  Télécharger XLSX fusionné",
                data=buf,
                file_name="mandats_fusionnes.xlsx",
                mime=XLSX_MIME,
                use_container_width=True,
            )

